import json
import openpyxl
import urllib.request
import urllib.parse
from datetime import datetime, date

from django.db import transaction
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib.auth.forms import UserCreationForm
from django.contrib.auth import authenticate, login, logout
from django.contrib import messages
from django.utils import timezone
from django.core.paginator import Paginator
from django.db.models import Q, Sum
from django.http import HttpResponse, JsonResponse
from django.utils.safestring import mark_safe
from django.core.mail import send_mail
from django.conf import settings

from .models import Uniform, Debt, Transaction, StaffDebt, ActionLog


# ==============================================================================
# HỆ THỐNG ĐĂNG NHẬP & ĐĂNG KÝ
# ==============================================================================

def xu_ly_dang_nhap(request):
    if request.user.is_authenticated:
        return redirect('dashboard')

    if request.method == 'POST':
        turnstile_response = request.POST.get('cf-turnstile-response')
        data = urllib.parse.urlencode({
            'secret': '0x4AAAAAADTcbRcgIPcBA48OWFqUfzxpEzA',  # Hãy chắc chắn đây là SECRET KEY
            'response': turnstile_response
        }).encode()

        try:
            req = urllib.request.Request('https://challenges.cloudflare.com/turnstile/v0/siteverify', data=data)
            response = urllib.request.urlopen(req)
            result = json.loads(response.read())
        except Exception:
            messages.error(request, "Không thể kết nối Cloudflare!")
            return redirect('login')

        if not result.get('success'):
            messages.error(request, "Xác thực Bot thất bại!")
            return redirect('login')

        tai_khoan = request.POST.get('username')
        mat_khau = request.POST.get('password')

        user = authenticate(request, username=tai_khoan, password=mat_khau)
        if user is not None:
            login(request, user)
            messages.success(request, f"Chào mừng {user.username} quay trở lại!")
            return redirect('dashboard')
        else:
            messages.error(request, "Tên đăng nhập hoặc mật khẩu không chính xác.")
            return redirect('login')

    return render(request, 'login.html')


def register(request):
    if request.method == 'POST':
        # Ép bộ form mặc định lấy thêm dữ liệu từ trường email
        form = UserCreationForm(request.POST)
        email_input = request.POST.get('email')  # Lấy email người dùng nhập từ giao diện

        if form.is_valid():
            user = form.save(commit=False)
            user.email = email_input  # Đẩy email vào tài khoản trước khi lưu chính thức
            user.save()

            # Tự động đăng nhập sau khi đăng ký thành công
            login(request, user)

            # --- Logic gửi email thông báo Admin (Giữ nguyên code cũ của bạn) ---

            try:
                tieu_de = f"🎉 [WMS] Người dùng mới: {user.username}"
                noi_dung = f"Tài khoản {user.username} vừa đăng ký lúc {timezone.now().strftime('%d/%m/%Y %H:%M:%S')}."
                nguoi_nhan = ['vinh81800135@gmail.com']
                send_mail(
                    subject=tieu_de, message=noi_dung,
                    from_email=settings.DEFAULT_FROM_EMAIL,
                    recipient_list=nguoi_nhan, fail_silently=True,
                )
            except Exception as e:
                print(f"Lỗi gửi email: {e}")

            messages.success(request, "Đăng ký thành công!")
            return redirect('dashboard')
    else:
        form = UserCreationForm()
    return render(request, 'register.html', {'form': form})


def xu_ly_dang_xuat(request):
    logout(request)
    return redirect('login')


# ==============================================================================
# DASHBOARD & THỐNG KÊ
# ==============================================================================

@login_required(login_url='login')
def dashboard(request):
    uniforms = Uniform.objects.all().order_by('name')
    low_stock_uniforms = Uniform.objects.filter(quantity__lt=10)

    top_stock = Uniform.objects.all().order_by('-quantity')[:7]
    chart_labels = [f"{u.name} ({u.size})" for u in top_stock]
    chart_data = [u.quantity for u in top_stock]

    return render(request, 'dashboard.html', {
        'uniforms': uniforms,
        'low_stock_uniforms': low_stock_uniforms,
        'low_stock_count': low_stock_uniforms.count(),
        'chart_labels_json': json.dumps(chart_labels),
        'chart_data_json': json.dumps(chart_data),
    })


@login_required(login_url='login')
def summary_nxt(request):
    uniforms = Uniform.objects.all()
    report_data = []
    for u in uniforms:
        nhap = Transaction.objects.filter(uniform=u, type='IN').aggregate(Sum('amount'))['amount__sum'] or 0
        xuat = Transaction.objects.filter(uniform=u, type='OUT').aggregate(Sum('amount'))['amount__sum'] or 0
        report_data.append({
            'sku': u.sku, 'name': u.name, 'size': u.size,
            'ton_dau': u.quantity - nhap + xuat, 'nhap': nhap, 'xuat': xuat, 'ton_cuoi': u.quantity,
        })
    return render(request, 'summary_nxt.html', {'report_data': report_data})


# ==============================================================================
# QUẢN LÝ KHO NHẬP / XUẤT (ĐÃ KHÓA ADMIN)
# ==============================================================================

@login_required(login_url='login')
def stock_action(request, action_type):
    if not request.user.is_superuser:
        messages.error(request, "Lỗi: Chỉ Admin mới có quyền Nhập/Xuất kho!")
        return redirect('dashboard')

    if request.method == 'POST':
        uniform_id = request.POST.get('uniform')
        amount = int(request.POST.get('amount', 1))
        uniform_obj = Uniform.objects.get(id=uniform_id)

        if action_type == 'OUT' and uniform_obj.quantity < amount:
            messages.error(request, f"Lỗi: Không đủ tồn kho!")
            return redirect('dashboard')

        Transaction.objects.create(
            ticket_id=request.POST.get('ticket_id'), uniform_id=uniform_id,
            amount=amount, type=action_type, actor_name=request.POST.get('actor_name')
        )

        uniform_obj.quantity += amount if action_type == 'IN' else -amount
        uniform_obj.save()

        ActionLog.objects.create(
            user=request.user, action="NHẬP KHO" if action_type == 'IN' else "XUẤT KHO",
            description=f"Đã {'nhập' if action_type == 'IN' else 'xuất'} {amount} '{uniform_obj.name}'."
        )
        return redirect('dashboard')

    template = 'add_stock.html' if action_type == 'IN' else 'issue_stock.html'
    return render(request, template, {'uniforms': Uniform.objects.all()})


@login_required(login_url='login')
def stock_form(request):
    if not request.user.is_superuser:
        messages.error(request, "Lỗi: Chỉ Admin mới có quyền Thao tác kho!")
        return redirect('dashboard')

    if request.method == 'POST':
        action_type = request.POST.get('type')
        uniform_id = request.POST.get('uniform')
        amount = int(request.POST.get('amount', 1))
        uniform_obj = Uniform.objects.get(id=uniform_id)

        if action_type == 'OUT' and uniform_obj.quantity < amount:
            messages.error(request, f"Lỗi: Không đủ tồn kho!")
            return redirect('dashboard')

        Transaction.objects.create(
            ticket_id=request.POST.get('ticket_id'), date=request.POST.get('date'),
            actor_name=request.POST.get('actor_name'), uniform_id=uniform_id,
            amount=amount, type=action_type, branch=request.POST.get('branch', 'Kho Tổng')
        )

        uniform_obj.quantity += amount if action_type == 'IN' else -amount
        uniform_obj.save()
        ActionLog.objects.create(
            user=request.user, action="NHẬP KHO" if action_type == 'IN' else "XUẤT KHO",
            description=f"Đã {'nhập' if action_type == 'IN' else 'xuất'} {amount} '{uniform_obj.name}'."
        )
        return redirect('dashboard')
    return render(request, 'stock_form.html', {'uniforms': Uniform.objects.all().order_by('name')})


# ==============================================================================
# NỢ HỌC SINH (DEBTS)
# ==============================================================================

@login_required(login_url='login')
def debt_list(request):
    # --- 1. XỬ LÝ LƯU KHOẢN NỢ MỚI (POST) ---
    if request.method == 'POST' and 'branch' in request.POST:
        uniform_id = request.POST.get('uniform')
        if uniform_id:
            Debt.objects.create(
                branch=request.POST.get('branch'),
                request_date=request.POST.get('request_date') or timezone.now().date(),
                uniform_id=uniform_id,
                quantity=int(request.POST.get('quantity', 1)),
                student_name=request.POST.get('student_name'),
                note="Kho nợ"
            )
            messages.success(request, "Lưu khoản nợ thành công!")
        return redirect('debt_list')

    # --- 2. XỬ LÝ BỘ LỌC ĐA NĂNG ĐỘC LẬP (GET FILTER) ---
    debts = Debt.objects.all().order_by('is_resolved', '-request_date')

    # Đọc dữ liệu từ 3 ô lọc trên giao diện (image_fcd046.png)
    search_query = request.GET.get('search', '').strip()
    filter_branch = request.GET.get('branch', '').strip()
    filter_status = request.GET.get('status', '').strip()

    # Ô 1: Tìm theo tên học sinh hoặc tên đồng phục
    if search_query:
        debts = debts.filter(
            Q(student_name__icontains=search_query) |
            Q(uniform__name__icontains=search_query)
        )

    # Ô 2: Lọc theo Cơ sở
    if filter_branch:
        debts = debts.filter(branch__icontains=filter_branch)

    # Ô 3: Lọc theo Trạng thái (Chưa trả hoặc Đã xong)
    if filter_status == 'resolved':
        debts = debts.filter(is_resolved=True)
    elif filter_status == 'pending':
        debts = debts.filter(is_resolved=False)

    # --- 3. TÍNH TOÁN SỐ LIỆU THỰC TẾ CHO 3 THẺ THỐNG KÊ ---
    total_debts = debts.aggregate(Sum('quantity'))['quantity__sum'] or 0  # Tổng giao dịch nợ
    pending_debts = debts.filter(is_resolved=False).aggregate(Sum('quantity'))['quantity__sum'] or 0  # Đang chờ trả mới
    resolved_debts = debts.filter(is_resolved=True).aggregate(Sum('quantity'))['quantity__sum'] or 0  # Đã hoàn thành

    # --- 4. PHÂN TRANG VÀ TRẢ DỮ LIỆU ---
    paginator = Paginator(debts, 15)
    return render(request, 'debt_list.html', {
        'page_obj': paginator.get_page(request.GET.get('page')),
        'uniforms': Uniform.objects.all(),
        # Giữ lại chữ trên các ô nhập sau khi bấm Lọc
        'search_query': search_query,
        'filter_branch': filter_branch,
        'filter_status': filter_status,
        # Truyền 3 con số thống kê chuẩn sang HTML
        'total_debts': total_debts,
        'pending_debts': pending_debts,
        'resolved_debts': resolved_debts,
    })


@login_required(login_url='login')
def resolve_debt(request, debt_id):
    if not request.user.is_superuser:
        messages.error(request, "Lỗi bảo mật: Chỉ Admin mới có quyền Gạch Nợ!")
        return redirect('debt_list')

    debt = get_object_or_404(Debt, id=debt_id)
    if not debt.is_resolved:
        if debt.uniform.quantity >= debt.quantity:
            debt.uniform.quantity -= debt.quantity
            debt.uniform.save()
            Transaction.objects.create(
                ticket_id=f"DEBT-{debt.id}", uniform=debt.uniform, amount=debt.quantity,
                type='OUT', actor_name="Hệ thống", branch=debt.branch
            )
            debt.is_resolved = True
            debt.note = f"{debt.note} (Đã trả {timezone.now().strftime('%d/%m/%Y')})"
            debt.save()

            ActionLog.objects.create(
                user=request.user, action="GẠCH NỢ HỌC SINH",
                description=f"Đã gạch nợ và xuất kho cho học sinh '{debt.student_name}'."
            )
            messages.success(request, f"Gạch nợ thành công cho {debt.student_name}!")
        else:
            messages.error(request, f"Không đủ tồn kho!")
    return redirect('debt_list')


@login_required(login_url='login')
def auto_resolve_debt(request, uniform_id):
    if not request.user.is_superuser:
        messages.error(request, "Lỗi bảo mật: Chỉ Admin mới có quyền Gạch Nợ Tự Động!")
        return redirect('debt_list')

    uniform = get_object_or_404(Uniform, id=uniform_id)
    pending_debts = Debt.objects.filter(uniform=uniform, is_resolved=False).order_by('request_date')

    resolved_count = 0
    for debt in pending_debts:
        if uniform.quantity >= debt.quantity:
            Transaction.objects.create(uniform=uniform, type='OUT', amount=debt.quantity, actor_name="Hệ thống",
                                       branch=debt.branch)
            uniform.quantity -= debt.quantity
            uniform.save()
            debt.is_resolved = True
            debt.save()
            resolved_count += 1

    if resolved_count > 0:
        ActionLog.objects.create(
            user=request.user, action="CẤN TRỪ TỰ ĐỘNG",
            description=f"Đã tự động cấn trừ {resolved_count} khoản nợ cho '{uniform.name}'."
        )
        messages.success(request, f"Đã cấn trừ {resolved_count} khoản nợ!")
    return redirect('debt_list')


@login_required(login_url='login')
def export_debt_excel(request):
    if not request.user.is_superuser:
        messages.error(request, "Lỗi bảo mật: Chỉ Admin mới có quyền Xuất Excel!")
        return redirect('debt_list')

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=Bao_cao_no.xlsx'
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.append(['Cơ sở', 'Ngày', 'Tên đồng phục', 'Size', 'SL', 'Học sinh', 'Ghi chú', 'Trạng thái'])

    for debt in Debt.objects.all().order_by('-request_date'):
        worksheet.append([debt.branch, str(debt.request_date), debt.uniform.name, debt.uniform.size, debt.quantity,
                          debt.student_name, debt.note, 'Đã xong' if debt.is_resolved else 'Chưa trả'])
    workbook.save(response)
    return response


@login_required(login_url='login')
def import_debt_excel(request):
    if not request.user.is_superuser:
        messages.error(request, "Lỗi bảo mật: Chỉ Admin mới có quyền Nhập Excel!")
        return redirect('debt_list')

    if request.method == 'POST' and request.FILES.get('excel_file'):
        # 🟢 1. BẮT TÍN HIỆU TỪ CHECKBOX XÓA DỮ LIỆU CŨ
        clear_data = request.POST.get('clear_old_data')
        if clear_data == 'yes':
            Debt.objects.all().delete()  # Xóa sạch toàn bộ dữ liệu nợ cũ

        try:
            wb = openpyxl.load_workbook(request.FILES['excel_file'], data_only=True)
            sheet = wb.active

            success_count = 0
            skip_count = 0  # Biến đếm số dòng bị trùng lặp

            for row in sheet.iter_rows(min_row=2, values_only=True):
                if not row[0] or not row[2]:
                    continue

                # Trích xuất dữ liệu
                branch_name = str(row[0]).strip()
                uniform_name = str(row[2]).strip()
                quantity_val = int(row[4] or 1)
                student_name = str(row[5]).strip() if row[5] else ""

                uniform = Uniform.objects.filter(name__icontains=uniform_name).first()

                if uniform:
                    # 🟢 2. LOGIC CHỐNG TRÙNG LẶP
                    # Kiểm tra xem Học sinh này ở Cơ sở này đã nợ Món đồ này chưa
                    is_exist = Debt.objects.filter(
                        branch=branch_name,
                        student_name=student_name,
                        uniform=uniform
                    ).exists()

                    if not is_exist:
                        # Nếu CHƯA CÓ -> Tạo mới
                        Debt.objects.create(
                            branch=branch_name,
                            request_date=timezone.now().date(),
                            uniform=uniform,
                            quantity=quantity_val,
                            student_name=student_name
                        )
                        success_count += 1
                    else:
                        # Nếu ĐÃ CÓ -> Bỏ qua, tăng biến đếm
                        skip_count += 1

            # 🟢 3. GHI LOG VÀ THÔNG BÁO THÔNG MINH
            if success_count > 0 or clear_data == 'yes':
                action_desc = f"Import {success_count} dòng nợ mới."
                if clear_data == 'yes':
                    action_desc = "XÓA DỮ LIỆU CŨ và " + action_desc

                ActionLog.objects.create(
                    user=request.user,
                    action="IMPORT NỢ HS",
                    description=action_desc
                )

            # Trả về thông báo cho người dùng
            if clear_data == 'yes':
                messages.success(request,
                                 f"Đã dọn sạch dữ liệu cũ! Import thành công {success_count} dòng mới. Bỏ qua {skip_count} dòng trùng.")
            else:
                messages.success(request,
                                 f"Nhập thành công {success_count} dòng! Bỏ qua {skip_count} dòng bị trùng lặp.")

        except Exception as e:
            messages.error(request, f"Lỗi đọc file: {str(e)}")

    return redirect('debt_list')


# ==============================================================================
# NỢ NHÂN VIÊN (STAFF DEBTS)
# ==============================================================================

@login_required(login_url='login')
def staff_debt_list(request):
    # --- 1. XỬ LÝ LƯU DỮ LIỆU CẤP PHÁT MỚI (POST) ---
    if request.method == 'POST':
        if not request.user.is_superuser:
            messages.error(request, "Lỗi bảo mật: Chỉ Admin mới có quyền Cấp Phát cho Nhân Viên!")
            return redirect('staff_debt_list')

        uniform_id = request.POST.get('uniform')
        if uniform_id:
            uniform_obj = Uniform.objects.get(id=uniform_id)
            StaffDebt.objects.create(
                employee_name=request.POST.get('employee_name'),
                position=request.POST.get('position', ''),  # Đã sửa thành position
                gender=request.POST.get('gender', 'Nam'),
                uniform=uniform_obj,
                quantity=int(request.POST.get('quantity', 1)),
                issue_date=request.POST.get('issue_date') or timezone.now().date(),
                branch=request.POST.get('branch', ''),
                note=request.POST.get('note', '')
            )
            ActionLog.objects.create(
                user=request.user, action="CẤP PHÁT NV",
                description=f"Cấp áo cho nhân viên {request.POST.get('employee_name')}."
            )
            messages.success(request, "Đã lưu dữ liệu cấp phát mới!")
        return redirect('staff_debt_list')

    # --- 2. XỬ LÝ BỘ LỌC ĐA NĂNG ĐỘC LẬP (GET FILTER) ---
    staff_debts = StaffDebt.objects.all()

    # Đọc dữ liệu từ 3 ô lọc trên giao diện (image_c82c27.png)
    search_query = request.GET.get('search', '').strip()
    filter_role = request.GET.get('position', '').strip()  # Nhận từ ô "Lọc theo Chức vụ" (Sửa lỗi sai tên biến role -> position)
    filter_branch = request.GET.get('branch', '').strip()  # Nhận từ ô "Lọc theo Cơ sở"

    # Ô 1: Tìm theo tên nhân viên hoặc tên sản phẩm đồng phục
    if search_query:
        staff_debts = staff_debts.filter(
            Q(employee_name__icontains=search_query) |
            Q(uniform__name__icontains=search_query)
        )

    # Ô 2: Tìm kiếm chức vụ (So khớp với trường position trong DB)
    if filter_role:
        staff_debts = staff_debts.filter(position__icontains=filter_role)

    # Ô 3: Tìm kiếm theo cơ sở
    if filter_branch:
        staff_debts = staff_debts.filter(branch__icontains=filter_branch)

    # Xử lý Sắp xếp (Sort) khi click vào tiêu đề
    sort_param = request.GET.get('sort', '').strip()
    valid_sort_fields = [
        'employee_name', '-employee_name', 'position', '-position', 
        'gender', '-gender', 'uniform__name', '-uniform__name', 
        'uniform__size', '-uniform__size', 'quantity', '-quantity', 
        'issue_date', '-issue_date', 'branch', '-branch', 
        'is_resolved', '-is_resolved'
    ]
    if sort_param in valid_sort_fields:
        staff_debts = staff_debts.order_by(sort_param)
    else:
        staff_debts = staff_debts.order_by('is_resolved', '-issue_date')

    tong_luot_cap_phat = staff_debts.aggregate(Sum('quantity'))['quantity__sum'] or 0

    # Đang sử dụng (Chưa trả) = các bản ghi có is_resolved=False
    dang_su_dung = staff_debts.filter(is_resolved=False).aggregate(Sum('quantity'))['quantity__sum'] or 0
    
    unique_positions = StaffDebt.objects.exclude(position='').values_list('position', flat=True).distinct()
    unique_branches = StaffDebt.objects.exclude(branch='').values_list('branch', flat=True).distinct()

    return render(request, 'staff_debt_list.html', {
        'page_obj': Paginator(staff_debts, 15).get_page(request.GET.get('page')),
        'uniforms': Uniform.objects.all().order_by('name'),
        # Giữ lại chữ trên các ô input sau khi trang web tải lại dữ liệu lọc
        'search_query': search_query,
        'filter_role': filter_role,
        'filter_branch': filter_branch,
        'sort_param': sort_param,
        'tong_luot_cap_phat': tong_luot_cap_phat,
        'dang_su_dung': dang_su_dung,
        'unique_positions': unique_positions,
        'unique_branches': unique_branches,
    })


@login_required(login_url='login')
def resolve_staff_debt(request, debt_id):
    if not request.user.is_superuser:
        messages.error(request, "Lỗi bảo mật: Tác vụ dành riêng cho Admin!")
        return redirect('staff_debt_list')

    debt = get_object_or_404(StaffDebt, id=debt_id)
    if not debt.is_resolved and debt.uniform.quantity >= debt.quantity:
        debt.is_resolved = True
        debt.uniform.quantity -= debt.quantity
        debt.uniform.save()
        Transaction.objects.create(
            ticket_id=f"STAFF-{debt.id}", uniform=debt.uniform, amount=debt.quantity,
            type='OUT', actor_name="Hệ thống", branch=debt.branch
        )
        debt.save()
        ActionLog.objects.create(user=request.user, action="GIAO ÁO NV",
                                 description=f"Nhân viên {debt.employee_name} đã nhận áo.")
        messages.success(request, "Đã giao áo!")
    return redirect('staff_debt_list')


@login_required(login_url='login')
def edit_staff_debt(request, debt_id):
    if not request.user.is_superuser:
        messages.error(request, "Lỗi bảo mật: Tác vụ dành riêng cho Admin!")
        return redirect('staff_debt_list')

    debt = get_object_or_404(StaffDebt, id=debt_id)

    if request.method == 'POST':
        debt.employee_name = request.POST.get('employee_name')
        debt.position = request.POST.get('position', '')  # Đã sửa thành position
        debt.gender = request.POST.get('gender', 'Nam')

        uniform_id = request.POST.get('uniform')
        if uniform_id:
            debt.uniform_id = uniform_id

        debt.quantity = int(request.POST.get('quantity', 1))

        # Xử lý cập nhật Ngày Xuất
        issue_date_str = request.POST.get('issue_date')
        if issue_date_str:
            debt.issue_date = issue_date_str

        # Xử lý cập nhật Ngày Trả (nếu có form điền vào)
        return_date_str = request.POST.get('return_date')
        if return_date_str:
            debt.return_date = return_date_str

        debt.branch = request.POST.get('branch', '')
        debt.note = request.POST.get('note', '')

        # Tự động cập nhật trạng thái nếu có tích chọn đã trả kho
        is_resolved_val = request.POST.get('is_resolved')
        if is_resolved_val is not None:
            debt.is_resolved = is_resolved_val == 'true' or is_resolved_val == 'on'

        debt.save()

        ActionLog.objects.create(
            user=request.user, action="SỬA THÔNG TIN NV",
            description=f"Cập nhật thông tin cấp phát của nhân viên {debt.employee_name}."
        )
        messages.success(request, f"Đã cập nhật thông tin thành công cho {debt.employee_name}!")
        return redirect('staff_debt_list')

    return redirect('staff_debt_list')

@login_required(login_url='login')
def delete_staff_debt(request, debt_id):
    if not request.user.is_superuser:
        messages.error(request, "Lỗi bảo mật: Tác vụ dành riêng cho Admin!")
        return redirect('staff_debt_list')

    debt = get_object_or_404(StaffDebt, id=debt_id)
    if debt.is_resolved:
        # Nếu đã cấp phát (đã trừ kho) thì nay xóa đi phải cộng trả lại kho
        debt.uniform.quantity += debt.quantity
        debt.uniform.save()
        Transaction.objects.create(
            ticket_id=f"DEL-STAFF-{debt.id}", uniform=debt.uniform, amount=debt.quantity,
            type='IN', actor_name="Hệ thống", branch=debt.branch
        )
    debt.delete()
    ActionLog.objects.create(user=request.user, action="XÓA DỮ LIỆU NV",
                             description=f"Đã xóa bản ghi của {debt.employee_name}.")
    messages.warning(request, "Đã xóa bản ghi!")
    return redirect('staff_debt_list')


@login_required(login_url='login')
def export_staff_debt_excel(request):
    """Xuất toàn bộ dữ liệu trong bảng cấp phát nhân viên ra file Excel"""
    if not request.user.is_superuser:
        messages.error(request, "Lỗi bảo mật: Chỉ Admin mới có quyền Xuất Excel!")
        return redirect('staff_debt_list')

    # 1. Khởi tạo một file Excel mới trên bộ nhớ đại diện
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=Bao_cao_cap_phat_nhan_vien.xlsx'

    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "Danh sách cấp phát"

    # 2. Tạo hàng tiêu đề (Header) cho các cột trong bảng dữ liệu
    worksheet.append([
        'STT', 'Tên nhân viên', 'Chức vụ', 'Giới tính',
        'Tên sản phẩm', 'Kích cỡ', 'Số lượng',
        'Ngày xuất', 'Ngày nhập (Trả lại)', 'Cơ sở', 'Ghi chú', 'Trạng thái'
    ])

    # 3. Lấy toàn bộ dữ liệu từ bảng StaffDebt xếp theo thứ tự ngày xuất mới nhất lên đầu
    staff_debts = StaffDebt.objects.all().order_by('is_resolved', '-issue_date')

    # 4. Vòng lặp đổ dữ liệu từng dòng vào file Excel
    for index, debt in enumerate(staff_debts, start=1):
        # Định dạng lại ngày tháng hiển thị dạng Ngày/Tháng/Năm cho dễ nhìn
        ngay_xuat = debt.issue_date.strftime('%d/%m/%Y') if debt.issue_date else ''
        ngay_nhap_lai = debt.return_date.strftime('%d/%m/%Y') if debt.return_date else '---'

        trang_thai = 'Đã nhận đủ' if debt.is_resolved else 'Đang sử dụng (Chưa trả)'

        worksheet.append([
            index,
            debt.employee_name,
            debt.position,
            debt.gender,
            debt.uniform.name,
            debt.uniform.size,
            debt.quantity,
            ngay_xuat,
            ngay_nhap_lai,
            debt.branch,
            debt.note or '',
            trang_thai
        ])

    # 5. Lưu cấu trúc và gửi file về cho trình duyệt của người dùng tự động tải xuống
    workbook.save(response)
    return response


@login_required(login_url='login')
def import_staff_debt_excel(request):
    """Đọc file Excel và import hàng loạt dữ liệu vào bảng cấp phát nhân viên"""
    if not request.user.is_superuser:
        messages.error(request, "Lỗi bảo mật: Chỉ Admin mới có quyền Nhập dữ liệu từ Excel!")
        return redirect('staff_debt_list')

    if request.method == 'POST' and request.FILES.get('excel_file'):
        # 🟢 1. BẮT TÍN HIỆU TỪ CHECKBOX XÓA DỮ LIỆU CŨ
        clear_data = request.POST.get('clear_old_data')
        if clear_data == 'yes':
            StaffDebt.objects.all().delete()  # Xóa sạch toàn bộ nợ nhân viên cũ

        try:
            excel_file = request.FILES['excel_file']

            # Đọc file excel từ bộ nhớ (data_only=True để lấy giá trị text thay vì công thức)
            wb = openpyxl.load_workbook(excel_file, data_only=True)
            sheet = wb.active

            success_count = 0
            error_count = 0

            from datetime import datetime
            
            def parse_excel_date(val):
                if not val or str(val).strip() in ['-', '---', '']:
                    return None
                if isinstance(val, datetime):
                    return val.date()
                val = str(val).strip()
                for fmt in ['%d/%m/%Y', '%Y-%m-%d', '%d-%m-%Y']:
                    try:
                        return datetime.strptime(val, fmt).date()
                    except ValueError:
                        pass
                return None

            # Vòng lặp đọc từ dòng thứ 2 (bỏ qua hàng tiêu đề)
            for row in sheet.iter_rows(min_row=2, values_only=True):
                # Nếu dòng đó trống rỗng (không có tên nhân viên) thì bỏ qua
                if not row[1]:
                    continue

                ten_nhan_vien = str(row[1]).strip()
                chuc_vu = str(row[2]).strip() if len(row) > 2 and row[2] else ''
                gioi_tinh = str(row[3]).strip() if len(row) > 3 and row[3] in ['Nam', 'Nữ'] else 'Nam'
                ten_san_pham = str(row[4]).strip() if len(row) > 4 and row[4] else ''
                kich_co = str(row[5]).strip() if len(row) > 5 and row[5] else 'Free Size'
                so_luong = int(row[6]) if len(row) > 6 and row[6] else 1
                
                ngay_xuat = row[7] if len(row) > 7 else None
                issue_date = parse_excel_date(ngay_xuat) or timezone.now().date()
                
                ngay_tra = row[8] if len(row) > 8 else None
                return_date = parse_excel_date(ngay_tra)
                
                co_so = str(row[9]).strip() if len(row) > 9 and row[9] else ''
                ghi_chu = str(row[10]).strip() if len(row) > 10 and row[10] else ''
                
                trang_thai = str(row[11]).strip().lower() if len(row) > 11 and row[11] else ''
                is_resolved = 'đủ' in trang_thai or 'trả' in trang_thai

                # Tìm kiếm sản phẩm trong kho bằng tên và kích cỡ
                uniform_obj = Uniform.objects.filter(name__icontains=ten_san_pham, size__iexact=kich_co).first()
                if not uniform_obj:
                    # Fallback tìm bằng tên nếu không khớp size
                    uniform_obj = Uniform.objects.filter(name__icontains=ten_san_pham).first()

                if uniform_obj:
                    # Tạo bản ghi cấp phát mới
                    StaffDebt.objects.create(
                        employee_name=ten_nhan_vien,
                        position=chuc_vu,
                        gender=gioi_tinh,
                        uniform=uniform_obj,
                        quantity=so_luong,
                        issue_date=issue_date,
                        return_date=return_date,
                        branch=co_so,
                        note=ghi_chu,
                        is_resolved=is_resolved
                    )
                    success_count += 1
                else:
                    # Ghi nhận số dòng lỗi do không tìm thấy tên áo tương ứng trong danh mục kho
                    error_count += 1

            # Lưu lại nhật ký hệ thống
            if success_count > 0:
                ActionLog.objects.create(
                    user=request.user,
                    action="IMPORT DATA NV",
                    description=f"Import thành công {success_count} nhân viên từ file Excel."
                )
                messages.success(request, f"Nhập thành công {success_count} dòng dữ liệu!")

            if error_count > 0:
                messages.warning(request, f"Có {error_count} dòng bị bỏ qua do tên sản phẩm không tồn tại trong kho.")

        except Exception as e:
            messages.error(request, f"Lỗi đọc file Excel: {str(e)}")

    return redirect('staff_debt_list')


# ==============================================================================
# NHẬT KÝ & IN ẤN (AUDIT LOG & RECEIPTS)
# ==============================================================================

@login_required(login_url='login')
def audit_log(request):
    if not request.user.is_superuser:
        messages.error(request, "Lỗi bảo mật: Bạn không có quyền xem Nhật ký hệ thống!")
        return redirect('dashboard')

    logs_list = ActionLog.objects.all().order_by('-created_at')
    return render(request, 'audit_log.html', {'page_obj': Paginator(logs_list, 30).get_page(request.GET.get('page'))})


from django.shortcuts import render


# Nhớ đảm bảo đã import StaffDebt ở đầu file nhé

@login_required
def print_receipt(request):
    ids_str = request.GET.get('ids', '')

    # Dùng Dictionary để gom nhóm dữ liệu theo Tên nhân viên
    grouped_items = {}


    if ids_str:
        id_list = ids_str.split(',')
        items = StaffDebt.objects.filter(id__in=id_list)

        if items.exists():
            co_so = items.first().branch

        # Bắt đầu vòng lặp gom nhóm
        for item in items:
            name = item.employee_name

            # Nếu tên nhân viên chưa có trong danh sách gom, thì tạo mới
            if name not in grouped_items:
                grouped_items[name] = {
                    'employee_name': name,
                    'position': item.position,
                    'qty_xanh': 0,
                    'qty_trang': 0,
                    'size': item.uniform.size,
                    'note': item.note or ""
                }
            else:
                # Nếu đã có tên rồi, kiểm tra xem size có khác không để gộp chuỗi (VD: "S, M")
                if item.uniform.size not in grouped_items[name]['size']:
                    grouped_items[name]['size'] += f", {item.uniform.size}"

                # Gộp ghi chú nếu có
                if item.note and item.note not in grouped_items[name]['note']:
                    grouped_items[name]['note'] += f" {item.note}"

            # Phân loại và cộng dồn số lượng áo dựa vào tên đồng phục
            uniform_name = item.uniform.name.lower()
            if 'xanh' in uniform_name:
                grouped_items[name]['qty_xanh'] += item.quantity
            elif 'trắng' in uniform_name or 'rap' in uniform_name:
                grouped_items[name]['qty_trang'] += item.quantity

    # Chuyển Dictionary thành dạng List (danh sách) để đẩy ra HTML dễ dàng
    final_items = list(grouped_items.values())

    return render(request, 'print_receipt.html', {
        'items': final_items,
        'co_so': co_so
    })


@login_required
def print_uniform_receipt(request):
    """ Hàm này CHỈ ĐỌC dữ liệu ra để hiển thị phôi xem trước, CHƯA TRỪ KHO """
    if request.method == 'POST':
        selected_ids = request.POST.getlist('uniform_ids')
        if not selected_ids:
            messages.error(request, "Vui lòng tích chọn ít nhất 1 mặt hàng!")
            return redirect('dashboard')

        selected_uniforms_data = []
        for u_id in selected_ids:
            item = get_object_or_404(Uniform, id=u_id)
            qty = int(request.POST.get(f'quantity_for_{u_id}', 1))

            selected_uniforms_data.append({
                'id': item.id,
                'name': item.name,
                'sku': item.sku if hasattr(item, 'sku') else item.id,
                'size': item.size if hasattr(item, 'size') else 'Tự do',
                'export_quantity': qty
            })

        context = {
            'selected_uniforms_data': selected_uniforms_data,
            'user': request.user,
        }
        return render(request, 'print_receipt_template.html', context)
    return redirect('dashboard')


@login_required
def confirm_export_stock(request):
    if request.method == 'POST':
        uniform_ids = request.POST.getlist('final_ids')
        quantities = request.POST.getlist('final_quantities')

        try:
            with transaction.atomic():
                audit_log = []  # Tên mảng mà Vinh đặt ở bước trước

                for idx, u_id in enumerate(uniform_ids):
                    item = get_object_or_404(Uniform, id=u_id)
                    qty = int(quantities[idx])

                    if item.quantity < qty:
                        # Thay vì redirect, ta trả về thông báo lỗi JSON
                        return JsonResponse(
                            {'status': 'error', 'message': f"Mặt hàng {item.name} không đủ tồn kho thực tế!"})

                    # Trừ kho
                    item.quantity -= qty
                    item.save()

                    audit_log.append(f"{item.name} (Size: {item.size}) [SL Xuất: {qty}]")

                # Ghi nhật ký thao tác
                action_detail = f"In phieu xuat {len(audit_log)} mat hang"

                # ÉP CẮT KHÓA CỨNG XUỐNG CẬN DƯỚI 45 KÝ TỰ ĐỂ BẢO ĐẢM AN TOÀN TUYỆT ĐỐI CHO MYSQL
                if len(action_detail) > 45:
                    action_detail = action_detail[:42] + "..."
                ActionLog.objects.create(
                    user=request.user,
                    action="Xuất Kho"
                )

                # 🟢 TRẢ VỀ KẾT QUẢ THÀNH CÔNG DẠNG JSON ĐỂ KÍCH HOẠT LỆNH IN
                return JsonResponse({'status': 'success'})

        except Exception as e:
            # 🔴 TRẢ VỀ LỖI HỆ THỐNG DẠNG JSON
            return JsonResponse({'status': 'error', 'message': str(e)})

    return redirect('dashboard')


@login_required(login_url='login')
def backup_database(request):
    """Xuất toàn bộ dữ liệu database dưới dạng JSON để backup"""
    if not request.user.is_superuser:
        messages.error(request, "Lỗi bảo mật: Chỉ Admin mới có quyền Sao lưu dữ liệu!")
        return redirect('dashboard')
        
    from django.core.management import call_command
    import io
    
    buf = io.StringIO()
    # Loại trừ contenttypes và auth.Permission để tránh lỗi khi restore
    call_command('dumpdata', natural_foreign=True, natural_primary=True, 
                 exclude=['contenttypes', 'auth.Permission'], indent=4, stdout=buf)
    
    buf.seek(0)
    response = HttpResponse(buf.read(), content_type='application/json')
    response['Content-Disposition'] = f'attachment; filename="wms_backup_{timezone.now().strftime("%Y%m%d_%H%M%S")}.json"'
    
    ActionLog.objects.create(
        user=request.user,
        action="SAO LƯU DỮ LIỆU",
        description="Tải xuống bản sao lưu toàn bộ hệ thống (JSON)"
    )
    return response